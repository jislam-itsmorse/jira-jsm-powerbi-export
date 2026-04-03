import os
import requests
import pandas as pd
from requests.auth import HTTPBasicAuth

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from datetime import datetime, timezone

# ==============================
# ENV VARIABLES
# ==============================
JIRA_BASE_URL = os.environ["JIRA_BASE_URL"]
JIRA_EMAIL = os.environ["JIRA_EMAIL"]
JIRA_API_TOKEN = os.environ["JIRA_API_TOKEN"]

SP_TENANT_ID = os.environ["SP_TENANT_ID"]
SP_CLIENT_ID = os.environ["SP_CLIENT_ID"]
SP_CLIENT_SECRET = os.environ["SP_CLIENT_SECRET"]

SP_SITE_HOSTNAME = os.environ["SP_SITE_HOSTNAME"]
SP_SITE_PATH = os.environ["SP_SITE_PATH"]
SP_LIBRARY_NAME = os.environ["SP_LIBRARY_NAME"]

# ==============================
# FILE CONFIG
# ==============================
LAST_SYNC_FILE = "last_sync.txt"
EXCEL_FILE = "jira_tickets.xlsx"

# ==============================
# JIRA CONFIG
# ==============================
FIELDS = [
    "summary",
    "status",
    "created",
    "updated",          # ✅ IMPORTANT
    "resolutiondate",
    "assignee",
    "issuetype"
]

# ==============================
# WATERMARK FUNCTIONS
# ==============================
def get_last_sync():
    if not os.path.exists(LAST_SYNC_FILE):
        return (datetime.now(timezone.utc) - pd.Timedelta(weeks=4)).isoformat()

    with open(LAST_SYNC_FILE, "r") as f:
        return f.read().strip()


def save_last_sync():
    now = datetime.now(timezone.utc).isoformat()
    with open(LAST_SYNC_FILE, "w") as f:
        f.write(now)

# ==============================
# JIRA FETCH
# ==============================
def fetch_jira_issues(jql):
    print("🔄 Fetching Jira issues...")
    print(jql.strip())

    all_issues = []
    next_page_token = None

    while True:
        params = {
            "jql": jql,
            "maxResults": 100,
            "fields": ",".join(FIELDS),
        }

        if next_page_token:
            params["nextPageToken"] = next_page_token

        response = requests.get(
            f"{JIRA_BASE_URL}/rest/api/3/search/jql",
            auth=HTTPBasicAuth(JIRA_EMAIL, JIRA_API_TOKEN),
            headers={"Accept": "application/json"},
            params=params,
        )

        print(f"Jira API Status: {response.status_code}")

        if response.status_code != 200:
            print(response.text)
            raise Exception("❌ Jira API call failed")

        data = response.json()
        issues = data.get("issues", [])

        all_issues.extend(issues)
        next_page_token = data.get("nextPageToken")

        if not next_page_token:
            break

    print(f"✅ Total issues fetched: {len(all_issues)}")
    return all_issues


# ==============================
# TRANSFORM
# ==============================
def issues_to_dataframe(issues):
    print("🔄 Transforming data...")

    rows = []
    for issue in issues:
        f = issue.get("fields", {}) or {}

        created = f.get("created")
        updated = f.get("updated")
        resolved = f.get("resolutiondate")

        rows.append({
            "IssueKey": issue.get("key"),
            "Summary": f.get("summary"),
            "Status": (f.get("status") or {}).get("name"),
            "CreatedDate": created,
            "UpdatedDate": updated,   # ✅ NEW
            "ResolvedDate": resolved,
            "Assignee": (f.get("assignee") or {}).get("displayName"),
            "IssueType": (f.get("issuetype") or {}).get("name"),
            "IsResolved": 1 if resolved else 0,
            "IsOpen": 1 if not resolved else 0
        })

    df = pd.DataFrame(rows)

    # ✅ Timezone-safe conversions
    df["CreatedDate"] = pd.to_datetime(df["CreatedDate"], errors="coerce", utc=True).dt.tz_convert(None)
    df["UpdatedDate"] = pd.to_datetime(df["UpdatedDate"], errors="coerce", utc=True).dt.tz_convert(None)
    df["ResolvedDate"] = pd.to_datetime(df["ResolvedDate"], errors="coerce", utc=True).dt.tz_convert(None)

    # ✅ Safety fallback
    df["UpdatedDate"] = df["UpdatedDate"].fillna(df["CreatedDate"])

    print(f"✅ Dataframe created with {len(df)} rows")
    return df


# ==============================
# GRAPH AUTH
# ==============================
def get_graph_token():
    url = f"https://login.microsoftonline.com/{SP_TENANT_ID}/oauth2/v2.0/token"

    response = requests.post(url, data={
        "client_id": SP_CLIENT_ID,
        "client_secret": SP_CLIENT_SECRET,
        "grant_type": "client_credentials",
        "scope": "https://graph.microsoft.com/.default",
    })

    if response.status_code != 200:
        print("❌ Token error:", response.text)
        raise Exception("Failed to get Graph token")

    return response.json()["access_token"]


def graph_get_site_id(token):
    url = f"https://graph.microsoft.com/v1.0/sites/{SP_SITE_HOSTNAME}:{SP_SITE_PATH}"

    response = requests.get(url, headers={"Authorization": f"Bearer {token}"})

    if response.status_code != 200:
        print("❌ Site lookup error:", response.text)
        raise Exception("Failed to resolve SharePoint site")

    return response.json()["id"]


def graph_get_drive_id(token, site_id):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"

    response = requests.get(url, headers={"Authorization": f"Bearer {token}"})

    if response.status_code != 200:
        print("❌ Drive lookup error:", response.text)
        raise Exception("Failed to list drives")

    for drive in response.json().get("value", []):
        if drive.get("name") == SP_LIBRARY_NAME:
            return drive["id"]

    raise Exception(f"❌ Drive not found: {SP_LIBRARY_NAME}")


def graph_upload_file(token, drive_id, local_path, target_name):
    print(f"⬆️ Uploading {target_name} to SharePoint...")

    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{target_name}:/content"

    with open(local_path, "rb") as f:
        response = requests.put(
            url,
            headers={"Authorization": f"Bearer {token}"},
            data=f
        )

    if response.status_code not in (200, 201):
        print("❌ Upload error:", response.text)
        raise Exception("Graph upload failed")

    print(f"✅ Uploaded: {target_name}")


# ==============================
# MAIN
# ==============================
if __name__ == "__main__":
    print("🚀 Starting Jira → SharePoint incremental export")

    # Step 1: Get last sync
    last_sync = get_last_sync()

    JIRA_QUERY_DYNAMIC = f"""
        project = ISD
        AND updated >= "{last_sync}"
        ORDER BY updated ASC
    """

    # Step 2: Fetch Jira data
    issues = fetch_jira_issues(JIRA_QUERY_DYNAMIC)

    # Step 3: Convert to DataFrame
    df = issues_to_dataframe(issues)

    # Step 4: Merge with existing file
    if os.path.exists(EXCEL_FILE):
        print("📂 Loading existing file for merge...")

        existing_df = pd.read_excel(EXCEL_FILE)

        existing_df["CreatedDate"] = pd.to_datetime(existing_df["CreatedDate"], errors="coerce")
        existing_df["UpdatedDate"] = pd.to_datetime(existing_df["UpdatedDate"], errors="coerce")
        existing_df["ResolvedDate"] = pd.to_datetime(existing_df["ResolvedDate"], errors="coerce")

        existing_df["UpdatedDate"] = existing_df["UpdatedDate"].fillna(existing_df["CreatedDate"])

        df = pd.concat([existing_df, df], ignore_index=True)

        # ✅ Correct deduplication
        df = df.sort_values("UpdatedDate").drop_duplicates(
            subset=["IssueKey"],
            keep="last"
        )

        print(f"🔁 After merge: {len(df)} records")

    if df.empty:
        print("⚠️ WARNING: No data to write")

    # Step 5: Save Excel
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Tickets")

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Tickets"]

    max_row = ws.max_row
    max_col = ws.max_column
    table_range = f"A1:{get_column_letter(max_col)}{max_row}"

    table = Table(displayName="TicketsTable", ref=table_range)

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(EXCEL_FILE)

    print(f"💾 Saved: {EXCEL_FILE}")

    # Step 6: Upload to SharePoint
    token = get_graph_token()
    site_id = graph_get_site_id(token)
    drive_id = graph_get_drive_id(token, site_id)

    graph_upload_file(token, drive_id, EXCEL_FILE, EXCEL_FILE)

    # Step 7: Save watermark
    save_last_sync()
    print("🕒 Sync timestamp updated")

    print("🎉 DONE")
